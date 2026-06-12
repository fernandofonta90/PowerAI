"""Lectura de archivos cargados (CSV y XLSX) a una tabla en memoria.

Devuelve una :class:`Tabla` con valores crudos en texto; la coerción de tipos y
la validación ocurren en capas posteriores.
"""

import csv
import io
from dataclasses import dataclass


class ArchivoIlegible(Exception):
    """El archivo no pudo leerse (formato no soportado o contenido corrupto)."""


@dataclass
class Tabla:
    """Tabla cruda: nombres de columna y filas como dicts de texto."""

    columnas: list[str]
    filas: list[dict[str, str]]


def _leer_csv(datos: bytes) -> Tabla:
    texto = datos.decode("utf-8-sig")
    lector = csv.reader(io.StringIO(texto))
    filas_brutas = [fila for fila in lector if any(c.strip() for c in fila)]
    if not filas_brutas:
        raise ArchivoIlegible("El archivo CSV está vacío.")
    encabezado = [c.strip() for c in filas_brutas[0]]
    filas: list[dict[str, str]] = []
    for cruda in filas_brutas[1:]:
        fila = {
            encabezado[i]: (cruda[i].strip() if i < len(cruda) else "")
            for i in range(len(encabezado))
        }
        filas.append(fila)
    return Tabla(columnas=encabezado, filas=filas)


def _leer_xlsx(datos: bytes) -> Tabla:
    from openpyxl import load_workbook

    libro = load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    hoja = libro.active
    if hoja is None:
        raise ArchivoIlegible("El archivo XLSX no tiene hojas.")

    iterador = hoja.iter_rows(values_only=True)
    try:
        encabezado_bruto = next(iterador)
    except StopIteration as exc:
        raise ArchivoIlegible("El archivo XLSX está vacío.") from exc

    encabezado = [str(c).strip() if c is not None else "" for c in encabezado_bruto]
    filas: list[dict[str, str]] = []
    for cruda in iterador:
        if cruda is None or all(c is None for c in cruda):
            continue
        fila = {
            encabezado[i]: ("" if i >= len(cruda) or cruda[i] is None else str(cruda[i]).strip())
            for i in range(len(encabezado))
        }
        filas.append(fila)
    libro.close()
    return Tabla(columnas=encabezado, filas=filas)


def leer_tabla(datos: bytes, nombre_archivo: str) -> Tabla:
    """Lee ``datos`` según la extensión de ``nombre_archivo`` (.csv o .xlsx)."""
    nombre = nombre_archivo.lower()
    if nombre.endswith(".csv"):
        return _leer_csv(datos)
    if nombre.endswith(".xlsx"):
        return _leer_xlsx(datos)
    raise ArchivoIlegible(f"Formato no soportado: {nombre_archivo}. Usa .csv o .xlsx.")
